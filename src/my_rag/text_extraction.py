import os
import fitz  # PyMuPDF
from docx import Document
import openpyxl
from pypdf import PdfReader

def extract_text_from_pdf(file_path: str) -> str:
    try:
        doc = fitz.open(file_path)
        text = "\n\n".join(page.get_text("text") for page in doc)
        doc.close()
        return text
    except Exception:
        reader = PdfReader(file_path)
        return "\n\n".join(page.extract_text() or "" for page in reader.pages)

def extract_text_from_docx(file_path: str) -> str:
    return "\n\n".join(p.text for p in Document(file_path).paragraphs)

def extract_text_from_xlsx(file_path: str) -> str:
    wb = openpyxl.load_workbook(file_path)
    parts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            parts.append(" | ".join(str(c) if c is not None else "" for c in row))
    return "\n\n".join(parts)

def extract_text_from_txt(file_path: str) -> str:
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()

def extract_text(file_path: str) -> str:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".pdf":
        return extract_text_from_pdf(file_path)
    if ext == ".docx":
        return extract_text_from_docx(file_path)
    if ext in [".xlsx", ".xls"]:
        return extract_text_from_xlsx(file_path)
    if ext == ".txt":
        return extract_text_from_txt(file_path)
    return f"[Không hỗ trợ định dạng: {ext}]"